"""
workers/feishu_docs.py · 飞书工作区读取 (0.9.0 BETA · 对标 OpenClaw feishu_doc/bitable/sheets)

能力:
  - parse_feishu_url(text)  识别消息文本里的飞书链接 (docx / sheets / base) → {type, token}
  - fetch_docx(token)       读云文档内容 (blocks 递归展开 → 纯文本)
  - fetch_sheet(token)      读电子表格 (sheet 列表 + 前 N 行值)
  - fetch_bitable(token)    读多维表格 (表列表 + 前 N 条记录)

权限 (开放平台 → 权限管理 → 添加后重新发布版本):
  - docx:document:readonly       读云文档
  - sheets:spreadsheet:readonly  读电子表格
  - bitable:app:readonly         读多维表格

错误约定: 403 = 权限未开 → 返回 ok=False + error 带 '权限' 字样提示用户去开放平台加权限。
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Optional

import requests

from workers.feishu_client import get_tenant_token

logger = logging.getLogger("opus.feishu.docs")

_BASE = "https://open.feishu.cn/open-apis"

# 限制拉取量 · 防大文档/大表把上下文撑爆
_MAX_BLOCKS = 200       # 文档最多展开 200 个 block
_MAX_TEXT_CHARS = 6000  # 文档文本上限
_MAX_SHEET_ROWS = 50    # 表格最多读 50 行
_MAX_SHEET_CHARS = 5000 # 表格文本上限
_MAX_BITABLE_ROWS = 50  # 多维表格最多读 50 条
_MAX_BITABLE_CHARS = 5000

# 飞书链接模式: https://xxx.feishu.cn/docx/AbCdEf123 ...
_LINK_RE = re.compile(r"https?://[a-z0-9.-]*feishu\.cn/(docx|sheets|base|wiki)/([A-Za-z0-9_\-]+)", re.I)


# ── URL 解析 ──────────────────────────────────────────────

def parse_feishu_url(text: str) -> Optional[dict]:
    """识别文本里的第一个飞书链接 → {type: docx|sheets|base|wiki, token}。"""
    if not text:
        return None
    m = _LINK_RE.search(text)
    if not m:
        return None
    return {"type": m.group(1).lower(), "token": m.group(2)}


def _api_get(path: str, params: dict | None = None, timeout: int = 15) -> dict:
    """带 token 的 GET · 返回飞书标准响应 {code, msg, data}。"""
    tok = get_tenant_token()
    if not tok:
        return {"ok": False, "error": "飞书未配置或 token 获取失败"}
    try:
        r = requests.get(
            f"{_BASE}{path}",
            headers={"Authorization": f"Bearer {tok}"},
            params=params or {},
            timeout=timeout,
        )
    except Exception as e:
        logger.warning("feishu docs api 异常: %s", e)
        return {"ok": False, "error": f"网络异常: {e}"}
    try:
        d = r.json()
    except Exception:
        # 非 JSON 响应 · 403 权限 / 400 校验失败常见
        if r.status_code == 403:
            return {"ok": False, "error": "权限不足 · 去开放平台加 docx/sheets/bitable 只读权限并重新发布版本"}
        if r.status_code == 400:
            return {"ok": False, "error": "链接 token 无效 (400) · 确认链接完整"}
        return {"ok": False, "error": f"飞书 API HTTP {r.status_code}"}
    if d.get("code") == 0:
        return {"ok": True, "data": d.get("data") or {}}
    code = d.get("code")
    msg = d.get("msg", "")
    # 99991663 / 99991672 常见权限类错误码 · 403 语义统一转成可读提示
    if code in (99991663, 99991672) or "permission" in msg.lower() or "权限" in msg:
        return {"ok": False, "error": "权限不足 · 去开放平台加 docx/sheets/bitable 只读权限并重新发布版本"}
    return {"ok": False, "error": f"飞书 API {code}: {msg}"}


# ── 云文档 docx ───────────────────────────────────────────

def _block_text(block: dict) -> str:
    """从单个 block 提取纯文本 (text / heading 等元素)。"""
    parts = []
    txt = block.get("text") or {}
    for el in txt.get("elements") or []:
        tr = el.get("text_run") or {}
        if tr.get("content"):
            parts.append(tr["content"])
        mr = el.get("mention_doc") or {}
        if mr.get("title"):
            parts.append(mr["title"])
    return "".join(parts).strip()


def _walk_blocks(blocks: list, out: list, depth: int = 0) -> None:
    """递归展开 block 树 (含 children)。"""
    if len(out) >= _MAX_BLOCKS:
        return
    for b in blocks:
        if len(out) >= _MAX_BLOCKS:
            return
        t = _block_text(b)
        if t:
            out.append(t)
        kids = b.get("children") or []
        if kids:
            _walk_blocks(kids, out, depth + 1)


def fetch_docx(token: str) -> dict:
    """读云文档 → 纯文本。"""
    meta = _api_get(f"/docx/v1/documents/{token}")
    if not meta["ok"]:
        return meta
    # blocks 接口返回树形结构 (含 children) · page_size 拉全
    blk = _api_get(f"/docx/v1/documents/{token}/blocks", {"page_size": 500})
    if not blk["ok"]:
        return blk
    items = blk["data"].get("items") or []
    lines = []
    _walk_blocks(items, lines)
    text = "\n".join(lines)[:_MAX_TEXT_CHARS]
    title = (meta["data"].get("document") or {}).get("title") or "未命名文档"
    return {"ok": True, "type": "docx", "title": title, "text": text, "chars": len(text)}


# ── 电子表格 sheets ───────────────────────────────────────

def fetch_sheet(token: str) -> dict:
    """读电子表格 → sheet 列表 + 第一个 sheet 的前 N 行。"""
    q = _api_get(f"/sheets/v3/spreadsheets/{token}/sheets/query")
    if not q["ok"]:
        return q
    sheets = q["data"].get("sheets") or []
    if not sheets:
        return {"ok": True, "type": "sheets", "title": "表格(空)", "text": "(空表格)"}
    first = sheets[0]
    sheet_id = first.get("sheet_id")
    title = first.get("title") or "Sheet1"
    # 读第一个 sheet 的值: range = sheet_id!A1:Z50 (上限 50 行)
    rng = f"{sheet_id}!A1:Z{_MAX_SHEET_ROWS}"
    vals = _api_get(f"/sheets/v3/spreadsheets/{token}/values/{rng}")
    if not vals["ok"]:
        # 列数可能超 Z · 退化成 A1:A50
        vals = _api_get(f"/sheets/v3/spreadsheets/{token}/values/{sheet_id}!A1:A{_MAX_SHEET_ROWS}")
        if not vals["ok"]:
            return vals
    value_range = vals["data"].get("value_range") or {}
    rows = value_range.get("values") or []
    lines = []
    for row in rows:
        cells = [str(c) if c is not None else "" for c in row]
        if any(cells):
            lines.append(" | ".join(cells))
    text = "\n".join(lines)[:_MAX_SHEET_CHARS]
    sheet_names = "、".join(s.get("title", "") for s in sheets[:10])
    return {
        "ok": True, "type": "sheets", "title": f"{title} (共{len(sheets)}个sheet: {sheet_names})",
        "text": text, "chars": len(text),
    }


# ── 多维表格 bitable ──────────────────────────────────────

def fetch_bitable(token: str) -> dict:
    """读多维表格 → 表列表 + 第一个表的前 N 条记录。"""
    t = _api_get(f"/bitable/v1/apps/{token}/tables", {"page_size": 50})
    if not t["ok"]:
        return t
    tables = t["data"].get("items") or []
    if not tables:
        return {"ok": True, "type": "bitable", "title": "多维表格(空)", "text": "(空表格)"}
    first = tables[0]
    table_id = first.get("table_id")
    table_name = first.get("name") or "表1"
    recs = _api_get(
        f"/bitable/v1/apps/{token}/tables/{table_id}/records",
        {"page_size": _MAX_BITABLE_ROWS},
    )
    lines = []
    if recs["ok"]:
        for r in recs["data"].get("items") or []:
            fields = r.get("fields") or {}
            parts = [f"{k}: {v}" for k, v in fields.items() if v not in (None, "")]
            if parts:
                lines.append(" · ".join(parts))
    text = "\n".join(lines)[:_MAX_BITABLE_CHARS]
    table_names = "、".join(x.get("name", "") for x in tables[:10])
    return {
        "ok": True, "type": "bitable", "title": f"{table_name} (共{len(tables)}个表: {table_names})",
        "text": text, "chars": len(text),
    }


# ── 消息文件读取 (0.9.0 BETA · 群里发的文件) ──────────────

_FILE_TEXT_EXTS = {".txt", ".md", ".log", ".csv", ".json", ".py", ".js", ".html"}
_FILE_MAX_BYTES = 20 * 1024 * 1024  # 20MB 上限


def fetch_file(message_id: str, file_key: str, file_name: str, max_chars: int = 6000) -> dict:
    """下载飞书消息里的文件并解析文本 (群里 @ 发文件场景)。

    下载: GET /im/v1/messages/{message_id}/resources/{file_key}?type=file
    解析: 文本类直读 · pdf (pypdf) · docx (python-docx) · xlsx (openpyxl)
    权限: im:resource (获取与上传图片或文件资源) · 未开返回可读提示。
    """
    if not file_key or not message_id:
        return {"ok": False, "error": "缺 file_key/message_id"}
    tok = get_tenant_token()
    if not tok:
        return {"ok": False, "error": "飞书未配置或 token 获取失败"}
    try:
        r = requests.get(
            f"{_BASE}/im/v1/messages/{message_id}/resources/{file_key}",
            params={"type": "file"},
            headers={"Authorization": f"Bearer {tok}"},
            timeout=30,
        )
    except Exception as e:
        logger.warning("feishu 文件下载异常: %s", e)
        return {"ok": False, "error": f"下载网络异常: {e}"}
    if r.status_code != 200:
        return {"ok": False, "error": f"下载失败 HTTP {r.status_code} · 需 im:resource 权限(去权限管理批量导入并发布)"}
    data = r.content
    if len(data) > _FILE_MAX_BYTES:
        return {"ok": True, "found": True, "type": "file", "title": file_name, "text": "", "unsupported": True,
                "error": f"文件 {len(data)//1024//1024}MB 超 20MB 上限"}
    ext = Path(file_name or "").suffix.lower()
    try:
        if ext in _FILE_TEXT_EXTS:
            text = data.decode("utf-8", errors="replace")
        elif ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(__import__("io").BytesIO(data))
            pages = [(p.extract_text() or "") for p in reader.pages[:10]]
            text = "\n".join(pages)
        elif ext == ".docx":
            from docx import Document
            doc = Document(__import__("io").BytesIO(data))
            text = "\n".join(p.text for p in doc.paragraphs if p.text)
        elif ext in (".xlsx", ".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(__import__("io").BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 50:
                    break
                vals = [str(c) for c in row if c is not None]
                if vals:
                    rows.append(" | ".join(vals))
            text = "\n".join(rows)
        else:
            return {"ok": True, "found": True, "type": "file", "title": file_name, "text": "",
                    "unsupported": True, "error": f"{ext} 格式暂不支持解析 (支持 txt/md/pdf/docx/xlsx)"}
    except Exception as e:
        logger.warning("feishu 文件解析异常: %s", e)
        return {"ok": False, "error": f"解析失败: {type(e).__name__}: {e}"}
    if not text or not text.strip():
        return {"ok": True, "found": True, "type": "file", "title": file_name, "text": "",
                "empty": True, "error": "文件无提取文本 (可能是扫描件/图片型 PDF)"}
    return {"ok": True, "found": True, "type": "file", "title": file_name, "text": text[:max_chars]}


# ── 统一入口 ──────────────────────────────────────────────

def fetch_image(message_id: str, image_key: str, save_dir: str = "data/cache/feishu_img") -> dict:
    """下载飞书消息里的图片存本地文件 (发图给 AI 看场景)。

    下载: GET /im/v1/messages/{message_id}/resources/{image_key}?type=image
      (消息资源接口 · 用户发来的图走这个 · im/v1/images/{key} 只能下机器人自己传的)
    权限: im:resource (获取与上传图片或文件资源) · 未开返回可读提示。
    返回: {"ok", "path", "ext"} · path 给 look_at 用。
    """
    if not image_key or not message_id:
        return {"ok": False, "error": "缺 image_key/message_id"}
    tok = get_tenant_token()
    if not tok:
        return {"ok": False, "error": "飞书未配置或 token 获取失败"}
    try:
        r = requests.get(
            f"{_BASE}/im/v1/messages/{message_id}/resources/{image_key}",
            params={"type": "image"},
            headers={"Authorization": f"Bearer {tok}"},
            timeout=30,
        )
    except Exception as e:
        logger.warning("feishu 图片下载异常: %s", e)
        return {"ok": False, "error": f"下载网络异常: {e}"}
    if r.status_code != 200:
        return {"ok": False, "error": f"下载失败 HTTP {r.status_code} · 需 im:resource 权限(去权限管理批量导入并发布)"}
    data = r.content
    if len(data) > _FILE_MAX_BYTES:
        return {"ok": False, "error": f"图片 {len(data)//1024//1024}MB 超 20MB 上限"}
    # 用 PIL 验证真实格式 → 按真实格式定扩展名 (content-type/magic 猜会漏 webp 等)
    ext = ".bin"
    try:
        from PIL import Image
        import io
        img = Image.open(io.BytesIO(data))
        img.load()  # 真解码 · 防"能打开但损坏"
        fmt = (img.format or "").upper()
        ext = {
            "PNG": ".png", "JPEG": ".jpg", "JPG": ".jpg", "GIF": ".gif",
            "WEBP": ".webp", "BMP": ".bmp",
        }.get(fmt, ".bin")
    except Exception as e:
        # 诊断: 飞书返回 200 但内容不是图片 (可能 JSON 错误/HTML) · dump 前 128 字节
        head = data[:128]
        logger.warning("feishu 图片解码失败: %s · ct=%s · head=%s",
                       e, r.headers.get("content-type"), head.hex())
        try:
            from pathlib import Path
            d = Path("data/cache/feishu_img")
            d.mkdir(parents=True, exist_ok=True)
            (d / "_debug_download.bin").write_bytes(data[:4096])
        except Exception:
            logger.warning("feishu_docs 解析异常失败 (L329)", exc_info=True)
        return {"ok": False, "error": f"图片解码失败: {type(e).__name__}: {e}"}
    if ext == ".bin":
        return {"ok": False, "error": f"不支持的图片格式: {img.format} · 支持 png/jpg/gif/webp/bmp"} if img else {"ok": False, "error": "无法识别图片格式"}
    try:
        from pathlib import Path
        d = Path(save_dir)
        d.mkdir(parents=True, exist_ok=True)
        p = d / f"{image_key.replace('/', '_')}{ext}"
        p.write_bytes(data)
    except Exception as e:
        logger.warning("feishu 图片存盘异常: %s", e)
        return {"ok": False, "error": f"存盘失败: {type(e).__name__}: {e}"}
    return {"ok": True, "path": str(p), "ext": ext, "size": len(data)}


def fetch_by_link(text: str) -> dict:
    """消息文本里若有飞书链接 → 自动识别并拉取内容。没有链接返回 None。"""
    parsed = parse_feishu_url(text)
    if not parsed:
        return {"ok": True, "found": False}
    t, token = parsed["type"], parsed["token"]
    try:
        if t == "docx":
            r = fetch_docx(token)
        elif t == "sheets":
            r = fetch_sheet(token)
        elif t == "base":
            r = fetch_bitable(token)
        elif t == "wiki":
            # wiki 是知识库页 · 先用 docx 试 (多数 wiki 页底层是 docx)
            r = fetch_docx(token)
            if not r["ok"]:
                r = {"ok": False, "error": f"wiki 页 {token} 读取失败 · 可能需先解析 obj_token (后续支持)"}
        else:
            r = {"ok": False, "error": f"暂不支持的飞书链接类型: {t}"}
        if r["ok"]:
            r["found"] = True
            r["url_type"] = t
        return r
    except Exception as e:
        logger.warning("feishu fetch_by_link 异常: %s", e)
        return {"ok": False, "error": f"读取失败: {e}"}
